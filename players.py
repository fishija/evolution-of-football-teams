import requests
from matches_data import barcelona_matches, real_madrid_matches, atlético_de_madrid_matches
from bs4 import BeautifulSoup
from openpyxl import Workbook



def search_for_players(ending_of_url, team1, team2, team_for_project):
    
    url = f'https://www.bdfutbol.com/en/p/p.php?id={ending_of_url}'
    response = requests.get(url)
    players = []

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        sections = soup.find_all('table', class_='taula_estil')
        
        for section_number in range(0,4):
            section=sections[section_number]
            rows = section.find_all('tr')[1:] 

            for row in rows:
                columns = row.find_all('td')
                player_name = columns[3].text

                if player_name!='' and team1==team_for_project:
                    if section_number==0 or section_number==2:
                        players.append(player_name)

                elif player_name!='' and team2==team_for_project:
                    if section_number==1 or section_number==3:
                        players.append(player_name)

    return players


#saving players  to excel
def save_to_excel(name_of_team, name_of_excel, matches_list):
    
    worbook = Workbook()
    sheet=worbook.active
    workbook_row=1

    for list_row in matches_list:
        try:
            if (list_row[5]== name_of_team and (list_row[6]>list_row[7])) or (list_row[9]==name_of_team and (list_row[7]>list_row[6])):
                players=search_for_players(list_row[8], list_row[5], list_row[9], name_of_team)

                sheet.cell(row=workbook_row, column=1, value=list_row[0])
                sheet.cell(row=workbook_row, column=2, value=list_row[5])
                sheet.cell(row=workbook_row, column=3, value=list_row[9])
                sheet.cell(row=workbook_row, column=4, value=f'{list_row[6]}-{list_row[7]}')
                print(list_row[0])
                for col, name in enumerate(players, start=5):
                    sheet.cell(row=workbook_row, column=col, value=name)
                
                workbook_row+=1
        except:
            worbook.save(f'{name_of_excel}.xlsx')

            worbook.close()
            quit()

    worbook.save(f'{name_of_excel}.xlsx')
    worbook.close()




save_to_excel('Real Madrid', 'Real_Madrid_players', real_madrid_matches)
save_to_excel('Barcelona', 'Barcelona_players', barcelona_matches)
save_to_excel('Atlético de Madrid', 'Atlético_de_Madrid', atlético_de_madrid_matches)
