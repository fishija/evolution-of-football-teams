import requests
from matches_data import (
    barcelona_matches,
    real_madrid_matches,
    atlético_de_madrid_matches,

    Barakaldo_matches,
    algeciras_matches,
    leganes_matches,
    mirandes_matches,
    girona,
)
from bs4 import BeautifulSoup
from openpyxl import Workbook


def search_for_players(ending_of_url, team1, team2, team_for_project):
    url = f"https://www.bdfutbol.com/en/p/p.php?id={ending_of_url}"
    response = requests.get(url)
    players = []

    if response.status_code == 200:
        soup = BeautifulSoup(response.text, "html.parser")
        sections = soup.find_all("table", class_="taula_estil")

        for section_number in range(0, 4):
            section = sections[section_number]
            rows = section.find_all("tr")[1:]

            for row in rows:
                columns = row.find_all("td")
                players_pseudonim = columns[3].text
                if players_pseudonim != "" and team1 == team_for_project:
                    if section_number == 0 or section_number == 2:
                        link = columns[3].find("a")  # Find the anchor tag in the column
                        try:
                            player_link = link.get(
                                "href"
                            )  # Get the href attribute of the anchor tag
                        except:
                            continue
                        url_for_one_player = f"https://www.bdfutbol.com/en/{player_link}"  # Construct the complete URL
                        response_for_players_name = requests.get(url_for_one_player)
                        soup_for_players = BeautifulSoup(
                            response_for_players_name.text, "html.parser"
                        )
                        sections_for_player = soup_for_players.find_all("title")
                        players_name = sections_for_player[0].text[
                            sections_for_player[0].text.find(",")
                            + 2 : sections_for_player[0].text.find(" -")
                        ]
                        players.append(players_name)

                elif players_pseudonim != "" and team2 == team_for_project:
                    if section_number == 1 or section_number == 3:
                        link = columns[3].find("a")  # Find the anchor tag in the column
                        try:
                            player_link = link.get(
                                "href"
                            )  # Get the href attribute of the anchor tag
                        except:
                            continue
                        url_for_one_player = f"https://www.bdfutbol.com/en/{player_link}"  # Construct the complete URL
                        response_for_players_name = requests.get(url_for_one_player)
                        soup_for_players = BeautifulSoup(
                            response_for_players_name.text, "html.parser"
                        )
                        sections_for_player = soup_for_players.find_all("title")
                        players_name = sections_for_player[0].text[
                            sections_for_player[0].text.find(",")
                            + 2 : sections_for_player[0].text.find(" -")
                        ]
                        print("name: ", players_name)
                        players.append(players_name)
    return players

def save_to_excel(name_of_team, name_of_excel, matches_list):
    for decade in range(1970, 2023, 10):
        worbook = Workbook()
        sheet = worbook.active
        sheet.cell(row=1, column=1, value="WinOrLost")
        sheet.cell(row=1, column=2, value="year")
        sheet.cell(row=1, column=3, value="team1")
        sheet.cell(row=1, column=4, value="team2")
        sheet.cell(row=1, column=5, value="score")
        sheet.cell(row=1, column=6, value="players")

        workbook_row = 2

        for list_row in matches_list:
            if decade <= list_row[0] < decade + 10:
                try:
                    if (list_row[5] == name_of_team and (list_row[6] > list_row[7])) or (
                            list_row[9] == name_of_team and (list_row[7] > list_row[6])
                    ):
                        sheet.cell(row=workbook_row, column=1, value=1)
                    else:
                        sheet.cell(row=workbook_row, column=1, value=0)

                    players = search_for_players(
                        list_row[8], list_row[5], list_row[9], name_of_team
                    )
                    sheet.cell(row=workbook_row, column=2, value=list_row[0])
                    sheet.cell(row=workbook_row, column=3, value=list_row[5])
                    sheet.cell(row=workbook_row, column=4, value=list_row[9])
                    sheet.cell(row=workbook_row, column=5, value=f"{list_row[6]}-{list_row[7]}")

                    for col, name in enumerate(players, start=6):
                        sheet.cell(row=workbook_row, column=col, value=name)

                    workbook_row += 1
                except Exception as e:
                    print(f"Error: {e}")

        if workbook_row > 2:  # If there's any data for the decade, save it
            try:
                worbook.save(f"./Football_team_evolution_csv/{name_of_excel}_{decade}-{decade+10}.csv")
            except Exception as e:
                print(f"Error while saving: {e}")
            finally:
                worbook.close()


save_to_excel("Algeciras", "Algeciras Players", algeciras_matches)

#save_to_excel("Real Madrid", "Real_Madrid_players", real_madrid_matches)
#save_to_excel("Barcelona", "Barcelona_players", barcelona_matches)
#save_to_excel("Atlético de Madrid", "Atlético_de_Madrid", atlético_de_madrid_matches)
# save_to_excel("Girona", "rl", girona)
# save_to_excel("Mirandés", "Mirandes Players", mirandes_matches)

print("Algeciras finished")
